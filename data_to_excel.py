def write_rows():
    from openpyxl import Workbook

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    # ws['A1'] = 42

    # Python types will automatically be converted
    import datetime
    # ws['A2'] = datetime.datetime.now()

    # Rows can also be appended
    ws.append([1, 2, 3])
    ws.append(['4', '5', '6', '7', datetime.datetime.now()])

    # Save the file
    wb.save("write_rows.xlsx")


def write_columns():
    import pandas as pd

    df = pd.DataFrame(
        {
            'Шапка таблицы 1': ['1 значение', '2 значение', '3 значение', '4 значение', '5 значение', '6 значение'],
            'Шапка таблицы 2': ['', '2 значение шапки 2', '', '', '', 'Italian Serie A (1)'],
            'Шапка таблицы 3': ['176000000', '188500000', 90000000, '100000000', '', '105000000']
        }
    )  # Длина списков должна быть одинаковой
    df.to_excel('write_columns.xlsx', index=False)


def main():
    write_rows()
    write_columns()


if __name__ == "__main__":
    main()
